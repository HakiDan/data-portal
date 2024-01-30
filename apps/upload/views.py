from django.shortcuts import render, HttpResponse, redirect
from django.views.decorators.csrf import csrf_protect
from django.utils.text import get_valid_filename
from openpyxl.utils import range_boundaries
from django.views.generic import View
from django.contrib import messages
from dateutil import relativedelta
from openpyxl import load_workbook
from django.conf import settings
from apps.upload.models import *
from core.tasks import *
from .forms import FileForm
from pathlib import Path
from .models import *
import datetime as dt
import pandas as pd
import numpy as np
import os, logging
import mimetypes
import os.path
import shutil
import csv
import re
from datetime import datetime, timedelta
from core.decorators import restrict_user
from django.utils.decorators import method_decorator
import logging
from django.contrib.auth.decorators import login_required
import collections
from pyzabbix import ZabbixMetric, ZabbixSender

mainPath = settings.MAINPATH
logger = logging.getLogger('main')
zabbix_sender = ZabbixSender(settings.ZABBIX_SERVER, settings.ZABBIX_PORT)
decorators_all = [login_required(login_url="/login/"), restrict_user("ia_group")]

def templateColumn(initiative):
    """_summary_
    Check data template based on masterlist

    Args:
        initiative (_type_): _description_

    Returns:
        _type_: _description_
    """    
    template = FileTemplate.objects.get(masterlistbudget__subinit_name=initiative)
    with template.file_template.open('r') as f:
        lines = f.readlines()
        lines = [s.replace('\n', '') for s in lines]
        headerLen = len(lines)
    return headerLen, lines


def check_init_id(sub_init_id, workingFile):
    """_summary_
    Check init_id from the file to match with the Masterlist
    
    Args:
        sub_init_id (_type_): _description_
        workingFile (_type_): _description_

    Returns:
        _type_: _description_
    """
    df = pd.read_csv(workingFile)
    if df['tarikh'].iloc[0] == 'yyyy-mm-dd':
        first_value = df['indeks_data'].iat[1]
    else:
        first_value = df['indeks_data'].iat[0]
        
    if first_value != sub_init_id:
        return True
    else:
        return False
    
    
def lowercase_header(workingFile):
    """_summary_
    Lowercase the columns (if dataframe columns are uppercase) & remove whitespace

    Args:
        workingFile (_type_): _description_
    """
    df = pd.read_csv(workingFile)
    df.columns = df.columns.str.lower()
    df.columns = df.columns.str.replace(' ', '')
    df.to_csv(workingFile, index=False)


def copy_file_to_target(src, dest, new_dest):
    """_summary_
    Copy file from /temp to /transition

    Args:
        src (_type_): _description_
        dest (_type_): _description_
        new_dest (_type_): _description_
    """
    if os.path.isfile(dest) == True:
        shutil.copy(src, new_dest)
    else:
        shutil.copy(src, dest)

       
def move_file_to_target_accepted(src, dest, new_dest):
    """_summary_
    Move file from /temp_download_accepted to /home_download_accepted (if file exists)

    Args:
        src (_type_): _description_
        dest (_type_): _description_
        new_dest (_type_): _description_

    Returns:
        _type_: _description_
    """
    if os.path.isfile(dest) == True:
        shutil.copy(src, new_dest)
        return Path(new_dest).name
    else:
        shutil.move(src, dest)
        return Path(dest).name
        
        
def move_file_to_target_rejected(src, dest, new_dest):
    """_summary_
    Move file from /temp_download_rejected to /home_download_rejected (if file exists)

    Args:
        src (_type_): _description_
        dest (_type_): _description_
        new_dest (_type_): _description_

    Returns:
        _type_: _description_
    """
    if os.path.isfile(dest) == True:
        shutil.copy(src, new_dest)
        return Path(new_dest).name
    else:
        shutil.move(src, dest)
        return Path(dest).name
        
        
def unmerge_cell_copy_top_value(file_name, file_ext, sheet_name):
    """_summary_
    Unmerged cell and copy the top value to the unmerged cells

    Args:
        file_name (_type_): _description_
        file_ext (_type_): _description_
        sheet_name (_type_): _description_
    """
    try:
        wbook = load_workbook("{}/temp/{}.{}".format(mainPath, file_name, file_ext))
        sheet = wbook[sheet_name]
        lookup = create_merged_cell_lookup(sheet)
        cell_group_list = lookup.keys()

        for cell_group in cell_group_list:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, 
                                       max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = lookup[cell_group]
        wbook.save("{}/temp/{}.xlsx".format(mainPath, file_name))

    except Exception as e:
        logging.exception("Exception occured")
        
        
def check_additional_col(workingFile, initiative):
    """_summary_
    Check additional columns for all templates

    Args:
        workingFile (_type_): _description_
        initiative (_type_): _description_

    Returns:
        _type_: _description_
    """
    df = pd.read_csv(workingFile)
    template = FileTemplate.objects.get(masterlistbudget__subinit_name=initiative)
    if template.template_name == 'financing':
        if ('kolum_tambahan_1' not in df and 'kolum_tambahan_2' not in df and 
            'kolum_tambahan_3' not in df and 'jumlah_rm_subsidi_disalurkan' not in df and 
            'jumlah_rm_subsidi_diluluskan' not in df):
            df['kolum_tambahan_1'] = 'NA'
            df['kolum_tambahan_2'] = 'NA'
            df['kolum_tambahan_3'] = 'NA'
            df['jumlah_rm_subsidi_disalurkan'] = 0
            df['jumlah_rm_subsidi_diluluskan'] = 0
        elif 'jumlah_rm_subsidi_disalurkan' not in df and 'jumlah_rm_subsidi_diluluskan' not in df:
            df['jumlah_rm_subsidi_disalurkan'] = 0
            df['jumlah_rm_subsidi_diluluskan'] = 0
        elif 'kolum_tambahan_1' not in df and 'kolum_tambahan_2' not in df and 'kolum_tambahan_3' not in df:
            df['kolum_tambahan_1'] = 'NA'
            df['kolum_tambahan_2'] = 'NA'
            df['kolum_tambahan_3'] = 'NA'
        elif 'kolum_tambahan_1' not in df:
            df['kolum_tambahan_1'] = 'NA'
        elif 'kolum_tambahan_2' not in df:
            df['kolum_tambahan_2'] = 'NA'
        elif 'kolum_tambahan_3' not in df:
            df['kolum_tambahan_3'] = 'NA'
        elif 'jumlah_rm_subsidi_diluluskan' not in df:
            df['jumlah_rm_subsidi_diluluskan'] = 'NA'
        elif 'jumlah_rm_subsidi_disalurkan' not in df:
            df['jumlah_rm_subsidi_disalurkan'] = 'NA'
        else:
            pass
    else:
        if 'kolum_tambahan_1' not in df and 'kolum_tambahan_2' not in df and 'kolum_tambahan_3' not in df:
            df['kolum_tambahan_1'] = 'NA'
            df['kolum_tambahan_2'] = 'NA'
            df['kolum_tambahan_3'] = 'NA'
        elif 'kolum_tambahan_1' not in df and 'kolum_tambahan_2' not in df:
            df['kolum_tambahan_1'] = 'NA'
            df['kolum_tambahan_2'] = 'NA'
        elif 'kolum_tambahan_1' not in df and 'kolum_tambahan_3' not in df:
            df['kolum_tambahan_1'] = 'NA'
            df['kolum_tambahan_3'] = 'NA'
        elif 'kolum_tambahan_2' not in df and 'kolum_tambahan_3' not in df:
            df['kolum_tambahan_2'] = 'NA'
            df['kolum_tambahan_3'] = 'NA'
        elif 'kolum_tambahan_1' not in df:
            df['kolum_tambahan_1'] = 'NA'
        elif 'kolum_tambahan_2' not in df:
            df['kolum_tambahan_2'] = 'NA'
        elif 'kolum_tambahan_3' not in df:
            df['kolum_tambahan_3'] = 'NA'
        else:
            pass
    df.to_csv(workingFile, mode='w+', index=False)
    return workingFile
    
def convert_to_csv(file_name, file_ext, sheet_name, initID, suffix):
    """_summary_
    Convert any text format to csv

    Args:
        file_name (_type_): _description_
        file_ext (_type_): _description_
        sheet_name (_type_): _description_
        initID (_type_): _description_
        suffix (_type_): _description_

    Returns:
        _type_: _description_
    """
    if file_ext == "csv":
        read_file = pd.read_csv(r"{}/temp/{}.csv".format(mainPath, file_name))
        src = "{}/temp/{}.csv".format(mainPath, file_name)
        dest = "{}/ori/{}/{}_{}.csv".format(mainPath, initID, file_name, suffix)
        shutil.copyfile(src, dest)
        return "{}/temp/{}.csv".format(mainPath, file_name)
        
    elif file_ext in ("xlsx", "xls"):
        read_file = pd.read_excel(r"{}/temp/{}.{}".format(mainPath, file_name, file_ext), sheet_name = sheet_name)
        unmerge_cell_copy_top_value(file_name, file_ext, sheet_name)
        num_rows = read_file.count()[0]
        # if(num_rows > 1000):
        #     sepRows(file_name, file_ext)
        read_file.to_csv(r"{}/temp/{}.csv".format(mainPath, file_name), index = False, header = True)
        read_file = read_file.fillna("NA")
        src = "{}/temp/{}.{}".format(mainPath, file_name, file_ext)
        dest = "{}/ori/{}/{}_{}.{}".format(mainPath, initID, file_name, suffix, file_ext)
        shutil.move(src, dest)
        return "{}/temp/{}.csv".format(mainPath, file_name)
    
    elif file_ext == "txt":
        read_file = list(csv.reader(open("{}/temp/{}.txt".format(mainPath, file_name), 'r'), delimiter='\t'))
        df = pd.DataFrame(read_file)
        df.columns = df.iloc[0]
        df = df[1:]
        df.to_csv("{}/temp/{}.csv".format(mainPath, file_name), index=None)
        return "{}/temp/{}.csv".format(mainPath, file_name)
    
    elif file_ext == "fwf":
        with open("{}/temp/{}.fwf".format(mainPath, file_name)) as infile:
            with open("{}/temp/{}.csv".format(mainPath, file_name), 'w') as outfile:
                data = infile.readlines()
                csv_data = [line.split() for line in data]
                for row in csv_data:
                    outfile.write(','.join(row) + '\n')
        return "{}/temp/{}.csv".format(mainPath, file_name)


def sepRows(file_name, file_ext): 
    """_summary_
    Seperate rows for file more than 5k

    Args:
        file_name (_type_): _description_
        file_ext (_type_): _description_
    """
    df = pd.read_excel("{}/temp/{}.{}".format(mainPath, file_name, file_ext)) 

    rows_per_file = 1000
    num_files = len(df) // rows_per_file + 1 

    dfs = [df[i:i+rows_per_file] for i in range(0, len(df), rows_per_file)]
    
    for i, df_chunk in enumerate(dfs):
        file_name = f'output_file_{i}.csv'
        df_chunk.to_csv(file_name, index=False)
    dfs = [pd.read_csv(f'output_file_{i}.csv') for i in range(num_files)]
    combined_df = pd.concat(dfs)
    combined_df.to_csv(r"{}/temp/{}.csv".format(mainPath, file_name), index = False, header = True)
    

def del_row_col(file_name):
    """_summary_
    Delete row(s) and column(s) not header(empty/title) 

    Args:
        file_name (_type_): _description_
    """
    read_file = pd.read_csv(r"{}/temp/{}.csv".format(mainPath, file_name))    
    read_file = read_file.dropna(axis=1, how='all')
    read_file = read_file.dropna(axis=0, how='all')
    new_header = read_file.iloc[0]
    read_file = read_file[1:]
    read_file.columns = new_header
    read_file.to_csv(r"{}/temp/{}.csv".format(mainPath, file_name), index = False, header = True)
    read_file = pd.read_csv(r"{}/temp/{}.csv".format(mainPath, file_name))
    num_rows = read_file.count()[0] 
    num_rows2 = read_file.shape[0]
    delRow = list(read_file[num_rows:].index)

    if num_rows2 > num_rows:
        for i in delRow:
            read_file.drop(i, inplace = True)
        read_file.to_csv(r"{}/temp/{}.csv".format(mainPath, file_name), index = False, header = True)
        read_file = pd.read_csv(r"{}/temp/{}.csv".format(mainPath, file_name))
       
  
def checkColumn(file_name, initiative):
    """_summary_
    Count the number of columns and return their status

    Args:
        request (_type_): _description_

    Returns:
        _type_: _description_
    """
    dataset = pd.read_csv(file_name)
    count_col = dataset.shape[1]
    listHeader = dataset.columns.values.tolist()
    if count_col < templateColumn(initiative)[0]:
        return '1', listHeader
    elif count_col > templateColumn(initiative)[0]:
        return '2', listHeader
    elif count_col == templateColumn(initiative)[0]:
        if collections.Counter(listHeader) == collections.Counter(templateColumn(initiative)[1]):
            return '3', listHeader
        else:
            return '2', listHeader
    
                    
def icValidation(icColumn, file_name, type_id):
    """_summary_
    IC validation after column checking

    Args:
        icColumn (_type_): _description_
        file_name (_type_): _description_

    Returns:
        _type_: _description_
    """
    ic_right = []
    ic_wrong = []
    all_ic = []
    type_right = []
    type_wrong = []
    type_id = type_id.astype({"jenis_id_ind":str})
    icColumn = icColumn.astype({"no_id_ind":str})
    ic_with_dash = '^[0-9]{6}\-[0-9]{2}\-[0-9]{4}$'
    ic_digits_only = '^[0-9]{9,12}$'
    old_ic_digits_only = '^[0-9]{6,8}$'
    old_ic_with_alpha = '^[a-zA-Z]{1,2}[0-9]{6,9}$'
    passport_1 = '^[a-zA-Z]{1}[0-9]{7}[a-zA-Z]{1}$'
    passport_2 = '^[a-zA-Z]{2}[0-9]{10}$'
    passport_3 = '^[0-9]{6}[a-zA-Z]{2}$'
    passport_4 = '^[a-zA-Z]{3}[0-9]{3}\/[0-9]{2}$'
    passport_5 = '^[a-zA-Z]{2}[0-9]{2}[a-zA-Z]{4}[0-9]{1}$' 

    id_lain = 'LAIN-LAIN'
    id_baru = 'IC BARU'
    id_lama = 'IC LAMA'

    for type_id, icColumn in zip(type_id.jenis_id_ind, icColumn.no_id_ind):
        if ((re.match(old_ic_with_alpha, icColumn) and re.match(type_id, id_lama)) or
            (re.match(old_ic_digits_only, icColumn) and re.match(type_id, id_lama)) or
            (re.match(ic_with_dash, icColumn) and re.match(type_id, id_baru)) or 
            (re.match(passport_1, icColumn) and re.match(type_id, id_lain)) or 
            (re.match(passport_2, icColumn) and re.match(type_id, id_lain)) or 
            (re.match(passport_3, icColumn) and re.match(type_id, id_lain)) or 
            (re.match(passport_4, icColumn) and re.match(type_id, id_lain)) or (re.match(passport_5,icColumn) and re.match(type_id, id_lain))):
            all_ic.append(icColumn)
            type_right.append(type_id)
            ic_right.append(icColumn)
        elif re.match(ic_digits_only, icColumn):
            icColumn = icColumn.zfill(12)
            icColumn = icColumn[:6] + '-' + icColumn[6:8] + '-' + icColumn[8:13]
            all_ic.append(icColumn)
            ic_right.append(icColumn)
        else:
            all_ic.append(icColumn)
            ic_wrong.append(icColumn)
            type_wrong.append(type_id)
    
    df = pd.read_csv("{}/temp/{}.csv".format(mainPath, file_name))
    df['no_id_ind'] = all_ic
    df['no_id_ind'] = df['no_id_ind'].replace('nan', '')
    df.to_csv("{}/temp/{}.csv".format(mainPath, file_name), mode='w+', index=False)

    if not ic_wrong:
        return "Success", ic_right, ic_wrong, type_right, type_wrong
    else:
        return "Failed", ic_right, ic_wrong, type_right, type_wrong


def brnValidation(brnColumn):
    """_summary_
    Business registration validation after column checking
    
    Args:
        brnColumn (_type_): _description_

    Returns:
        _type_: _description_
    """
    brn_right = []
    brn_wrong = []
    brnColumn = brnColumn.astype({"no_brn_entiti":str})
    invalid_value = ["NA", "nan", None, "NULL", "N/A"]
    
    for row in brnColumn.no_brn_entiti:
        if row:
            if row not in invalid_value:
                brn_right.append(row)
            else:
                brn_wrong.append(row)
        else:
            brn_wrong.append(row)

    if brn_right:
        return "Success", brn_right, brn_wrong
    else:
        return "Failed", brn_right, brn_wrong

  
def dateValidation(dateColumn):
    """_summary_
    Date validation after column checking

    Args:
        dateColumn (_type_): _description_

    Returns:
        _type_: _description_
    """
    date_wrong = []
    date_right = []
    date_format_dash_time = '^[0-9]{4}\-[0-9]{1,2}\-[0-9]{1,2}\ [0-9]{1,2}\:[0-9]{1,2}\:[0-9]{1,2}$'
    date_format_dash = '^[0-9]{4}\-[0-9]{1,2}\-[0-9]{1,2}$'
    date_ex = 'yyyy-mm-dd'
    dateColumn = dateColumn.fillna('NA')

    now = datetime.now()
    for row in dateColumn.tarikh:
        row = row.replace(' ', '')
        if (re.match(date_format_dash, row) or row == date_ex or 
            re.match(date_format_dash_time, row)):
            datetime_object = datetime.strptime(row, '%Y-%m-%d')
            if datetime_object > now:
                date_wrong.append(row)
            else:
                date_right.append(row)
        else:
            date_wrong.append(row)
    
    if not date_wrong:
        return "Success", date_right, date_wrong
    else:
        return "Failed", date_right, date_wrong
    
    
def check_mand_col(dataframe, initiative):
    """_summary_
    Check mandatory columns name for detail and summary data.

    Args:
        dataframe (_type_): _description_
        initiative (_type_): _description_

    Returns:
        _type_: _description_
    """
    template_detail = ('employment', 'program', 'ind_biz_assistance')
    template_summary = ('summary_a', 'summary_b')
    template = FileTemplate.objects.get(masterlistbudget__subinit_name=initiative)
    
    if template.template_name == 'financing':
        if 'no_brn_entiti' in dataframe.columns and 'tarikh' in dataframe.columns:
            return 'Passed'
        else:
            return 'Failed Financing'
          
    elif template.template_name in template_detail:
        if ('no_brn_entiti' in dataframe.columns and 'tarikh' in dataframe.columns and 
            'no_id_ind' in dataframe.columns):
            return 'Passed'
        else:
            return 'Failed Detail'
        
    elif template.template_name in template_summary:
        if 'tarikh' in dataframe.columns:
            return 'Passed'
        else:
            return 'Failed Summary'

def categorise(row):
    """_summary_
    Returns beneficiary category based on column no_brn_entiti

    Args:
        row (_type_): _description_

    Returns:
        _type_: _description_
    """
    if row['no_id_ind'] != 'nan' and row['no_brn_entiti'] != 'nan':
        return 'INDIVIDU'
    elif row['no_id_ind'] != 'nan':
        return 'INDIVIDU'
    else:
        return 'ENTITI'
            
            
def procSpecificCol(file_name, sub_init_id):
    """_summary_
    Process the file in /temp

    Args:
        file_name (_type_): _description_
        sub_init_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    dataset = pd.read_csv("{}/temp/{}.csv".format(mainPath, file_name))
    dataset_wo_headers = pd.read_csv("{}/temp/{}.csv".format(mainPath, file_name), header=None)
    list_col_name = list(dataset.columns.values)
    countCol = len(list_col_name)
    icField = dataset.filter(["no_id_ind"])
    brnField = dataset.filter(["no_brn_entiti"])
    dateField = dataset.filter(["tarikh"])
    typeField = dataset.filter(["jenis_id_ind"])
    
    if 'no_brn_entiti' in dataset.columns:
        if 'no_id_ind' in dataset.columns:
            dataset.no_brn_entiti = dataset.no_brn_entiti.astype(str)
            dataset.no_id_ind = dataset.no_id_ind.astype(str)
            resultright = (dataset[dataset['tarikh'].isin(dateValidation(dateField)[1]) & 
                                   ((dataset['no_brn_entiti'].isin(brnValidation(brnField)[1])) | 
                                    (dataset['no_id_ind'].isin(icValidation(icField, file_name, typeField)[1])))])
            resultwrong = dataset[~dataset.apply(tuple,1).isin(resultright.apply(tuple,1))]
            resultwrong['indeks_data'] = sub_init_id
            resultright['indeks_data'] = sub_init_id
            resultwrong = resultwrong.drop(resultwrong[resultwrong.tarikh == 'yyyy-mm-dd'].index)
            resultright = resultright.drop(resultright[resultright.tarikh == 'yyyy-mm-dd'].index)
            resultwrong_unique = resultwrong.drop_duplicates(subset=['no_id_ind', 'no_brn_entiti'])
            resultright_unique = resultright.drop_duplicates(subset=['no_id_ind', 'no_brn_entiti'])
            count_wrong = resultwrong_unique.no_id_ind.unique()
            count_right = resultright_unique.no_id_ind.unique()
            count_wrong = len([x for x in count_wrong if str(x) != 'nan'])
            count_right = len([x for x in count_right if str(x) != 'nan'])
            count_row_rejected = len(resultwrong.index)
            count_row_accepted = len(resultright.index)
            resultwrong_show = resultwrong.head(50)
            resultright_show = resultright.head(50)
            total_rejected_disb = resultwrong['jumlah_rm_disalurkan'].sum()
            total_accepted_disb = resultright['jumlah_rm_disalurkan'].sum()
            total_rejected_disb = format(total_rejected_disb, ',.2f')
            total_accepted_disb = format(total_accepted_disb, ',.2f')
            if 'kategori_penerima' in dataset.columns:
                resultright['kategori_penerima'] = resultright.apply(lambda resultright: categorise(resultright), axis=1)
                resultright_show['kategori_penerima'] = resultright_show.apply(lambda resultright_show: 
                                                        categorise(resultright_show), axis=1)
        else:
            dataset.no_brn_entiti = dataset.no_brn_entiti.astype(str)
            resultright = (dataset[dataset['tarikh'].isin(dateValidation(dateField)[1]) & 
                                   (dataset['no_brn_entiti'].isin(brnValidation(brnField)[1]))])
            resultwrong = dataset[~dataset.apply(tuple,1).isin(resultright.apply(tuple,1))]
            resultwrong['indeks_data'] = sub_init_id
            resultright['indeks_data'] = sub_init_id
            resultwrong = resultwrong.drop(resultwrong[resultwrong.tarikh == 'yyyy-mm-dd'].index)
            resultright = resultright.drop(resultright[resultright.tarikh == 'yyyy-mm-dd'].index)
            resultwrong_unique = resultwrong.drop_duplicates(subset=['no_brn_entiti'])
            resultright_unique = resultright.drop_duplicates(subset=['no_brn_entiti'])
            count_wrong = resultwrong_unique.no_brn_entiti.unique()
            count_right = resultright_unique.no_brn_entiti.unique()
            count_wrong = len([x for x in count_wrong if str(x) != 'nan'])
            count_right = len([x for x in count_right if str(x) != 'nan'])
            count_row_rejected = len(resultwrong.index)
            count_row_accepted = len(resultright.index)
            resultwrong_show = resultwrong.head(50)
            resultright_show = resultright.head(50)
            total_rejected_disb = resultwrong['jumlah_rm_disalurkan'].sum()
            total_accepted_disb = resultright['jumlah_rm_disalurkan'].sum()
            total_rejected_disb = format(total_rejected_disb, ',.2f')
            total_accepted_disb = format(total_accepted_disb, ',.2f')
            if 'kategori_penerima' in dataset.columns:
                resultright['kategori_penerima'] = resultright.apply(lambda resultright: categorise(resultright), axis=1)
                resultright_show['kategori_penerima'] = resultright_show.apply(lambda resultright_show: 
                                                        categorise(resultright_show), axis=1)
    else:
        if not dataset['negeri'].isnull().all():
                dataset['negeri'] = dataset['negeri'].str.upper()
        resultright = dataset[(dataset['tarikh'].isin(dateValidation(dateField)[1]))]
        resultwrong = dataset[~dataset.apply(tuple,1).isin(resultright.apply(tuple,1))]
        resultwrong['indeks_data'] = sub_init_id
        resultright['indeks_data'] = sub_init_id
        resultwrong = resultwrong.drop(resultwrong[resultwrong.tarikh == 'yyyy-mm-dd'].index)
        resultright = resultright.drop(resultright[resultright.tarikh == 'yyyy-mm-dd'].index)
        count_wrong = len(resultwrong[resultwrong.columns].drop_duplicates())
        count_right = len(resultright[resultright.columns].drop_duplicates())
        count_row_rejected = len(resultwrong.index)
        count_row_accepted = len(resultright.index)
        resultwrong_show = resultwrong.head(50)
        resultright_show = resultright.head(50)
        if 'jumlah_rm_disalurkan' in dataset.columns:
            total_rejected_disb = resultwrong['jumlah_rm_disalurkan'].sum()
            total_accepted_disb = resultright['jumlah_rm_disalurkan'].sum()
            total_rejected_disb = format(total_rejected_disb, ',.2f')
            total_accepted_disb = format(total_accepted_disb, ',.2f')
        else:
            total_rejected_disb = (resultwrong['jumlah_rm_disalurkan_1'].sum() + 
                                   resultwrong['jumlah_rm_disalurkan_2'].sum() + 
                                   resultwrong['jumlah_rm_disalurkan_3'].sum() + 
                                   resultwrong['jumlah_rm_disalurkan_4'].sum())
            total_accepted_disb = (resultright['jumlah_rm_disalurkan_1'].sum() + 
                                   resultright['jumlah_rm_disalurkan_2'].sum() + 
                                   resultright['jumlah_rm_disalurkan_3'].sum() + 
                                   resultright['jumlah_rm_disalurkan_4'].sum())
            total_rejected_disb = format(total_rejected_disb, ',.2f')
            total_accepted_disb = format(total_accepted_disb, ',.2f')
    
    resultwrong = resultwrong.replace('nan', '')
    resultright = resultright.replace('nan', '')
    resultwrong.to_csv("{}/temp_download_rejected/rejected_rows_{}.csv".format(mainPath, file_name), index=False)
    resultright.to_csv("{}/temp_download_accepted/accepted_rows_{}.csv".format(mainPath, file_name), index=False)
    resultwrong_filename = "rejected_rows_{}.csv".format(file_name).rsplit(".", 1)[0]
    resultright_filename = "accepted_rows_{}.csv".format(file_name).rsplit(".", 1)[0]
    result_all = {
        'list_col_name':list_col_name, 'icField':icField, 'dataset':dataset,
        'dataset_wo_headers':dataset_wo_headers, 'resultwrong_show':resultwrong_show,
        'resultright_show':resultright_show, 'total_rejected_disb':total_rejected_disb,
        'total_accepted_disb':total_accepted_disb, 'resultwrong_filename':resultwrong_filename,
        'resultright_filename':resultright_filename, 'count_wrong':count_wrong, 'count_right':count_right,
        'countCol':countCol, 'count_row_rejected':count_row_rejected, 'count_row_accepted':count_row_accepted,
        'brnField':brnField, 'dateField':dateField, 'resultwrong':resultwrong, 'resultright':resultright,
    }
    return result_all
    

@csrf_protect
def replaceHeader(request):
    """_summary_
    Replace IA header with Template header

    Args:
        request (_type_): _description_

    Returns:
        _type_: _description_
    """
    if request.method == 'POST':
        file_name = request.POST['file_name']
        iaHeader = request.POST['IAHeader']
        selectedHeader = request.POST['selectedHeader']
        correctionData = pd.read_csv("{}/temp/{}.csv".format(mainPath, file_name))
        dict = {iaHeader:selectedHeader}
        correctionData.rename(columns=dict, inplace=True)
        correctionData.to_csv("{}/temp/{}.csv".format(mainPath, file_name), index=False)
        return HttpResponse(status=200)
    return HttpResponse(status=400)
     
     
@csrf_protect
def cancel_submission(request):
    """_summary_
    Remove working file from /temp folder

    Args:
        request (_type_): _description_

    Returns:
        _type_: _description_
    """
    if request.method == 'POST':
        file_name = request.POST['file_name']
        initiative = request.POST['initiative']
        os.remove(r"{}/temp/{}.csv".format(mainPath, file_name))
        os.remove(r"{}/temp_download_accepted/accepted_rows_{}.csv".format(mainPath, file_name))
        os.remove(r"{}/temp_download_rejected/rejected_rows_{}.csv".format(mainPath, file_name))
        logger.info("Cancel file " + file_name + ".csv - Initiative: " + 
                    str(initiative) + " - " + str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Cancel file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                                   str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        return HttpResponse(status=200)
    logger.info("Failed cancel file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                str(request.user.username))
    try:
        packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                               "Failed cancel file " + file_name + ".csv - Initiative: " + 
                               str(initiative) + " - " + str(request.user.username))]
        zabbix_sender.send(packet)
    except:
        pass
    return HttpResponse(status=400)


def usersData(file_name, sub_init_id):
    """_summary_
    Return cleaned Dataframe and list of dictionaries after map to the mapping table

    Args:
        file_name (_type_): _description_
        sub_init_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    dataset = pd.read_csv(r"{}/temp/{}.csv".format(mainPath, file_name))
    
    df_wrong = procSpecificCol(file_name, sub_init_id)['resultwrong']
    df_right = procSpecificCol(file_name, sub_init_id)['resultright']
    
    gender = MapTableGender.objects.values_list('gender', flat=True)
    std_gender = MapTableGender.objects.values_list('std_gender', flat=True)
    dict_gender = dict(zip(gender, std_gender))
    
    race = MapTableRace.objects.values_list('race', flat=True)
    std_race = MapTableRace.objects.values_list('std_race', flat=True)
    dict_race = dict(zip(race, std_race))
    
    state = MapTableState.objects.values_list('state', flat=True)
    std_state = MapTableState.objects.values_list('std_state', flat=True)
    dict_state = dict(zip(state, std_state))
    
    bumi_status = MapTableBumiStatus.objects.values_list('bumi_status', flat=True)
    std_bumi_status = MapTableBumiStatus.objects.values_list('std_bumi_status', flat=True)
    dict_bumi_status = dict(zip(bumi_status, std_bumi_status))
    
    biz_sector = MapTableBizSector.objects.values_list('biz_sector', flat=True)
    std_biz_sector = MapTableBizSector.objects.values_list('std_biz_sector', flat=True)
    dict_biz_sector = dict(zip(biz_sector, std_biz_sector))
    
    biz_ownership_type = MapTableBizOwnershipType.objects.values_list('biz_ownership_type', flat=True)
    std_biz_ownership_type = MapTableBizOwnershipType.objects.values_list('std_biz_ownership_type', flat=True)
    dict_biz_ownership_type = dict(zip(biz_ownership_type, std_biz_ownership_type))

    df = pd.DataFrame(dataset)
    
    df_wrong = df_wrong.replace({"jantina_ind":dict_gender, "bangsa_ind":dict_race, 
                                 "negeri_ind":dict_state, "negeri":dict_state, 
                                 "status_bumiputera_entiti":dict_bumi_status, 
                                 "sektor_entiti":dict_biz_sector, 
                                 "jenis_entiti":dict_biz_ownership_type})
    df_right = df_right.replace({"jantina_ind":dict_gender, "bangsa_ind":dict_race, 
                                 "negeri_ind":dict_state, "negeri":dict_state, 
                                 "status_bumiputera_entiti":dict_bumi_status, 
                                 "sektor_entiti":dict_biz_sector, 
                                 "jenis_entiti":dict_biz_ownership_type})

    if 'jantina_ind' in df_right.columns:
        df_right['jantina_ind'] = df_right['jantina_ind'].astype(str)
        for data in df_right['jantina_ind'].unique():
            if data.upper() not in gender:
                MapTableGender.objects.get_or_create(gender=data.upper(), 
                                                     std_gender=data.upper(), 
                                                     sub_init_id=sub_init_id)
    
    if 'bangsa_ind' in df_right.columns:
        df_right['bangsa_ind'] = df_right['bangsa_ind'].astype(str)
        for data in df_right['bangsa_ind'].unique():
            if data.upper() not in race:
                MapTableRace.objects.get_or_create(race=data.upper(), 
                                                   std_race=data.upper(), 
                                                   sub_init_id=sub_init_id)
    
    if 'negeri_ind' in df_right.columns:
        df_right['negeri_ind'] = df_right['negeri_ind'].astype(str)
        for data in df_right['negeri_ind'].unique():
            if data.upper() not in state:
                MapTableState.objects.get_or_create(state=data.upper(), 
                                                    std_state=data.upper(), 
                                                    sub_init_id=sub_init_id)
                
    if 'negeri' in df_right.columns:
        df_right['negeri'] = df_right['negeri'].astype(str)
        for data in df_right['negeri'].unique():
            if data.upper() not in state:
                MapTableState.objects.get_or_create(state=data.upper(), 
                                                    std_state=data.upper(), 
                                                    sub_init_id=sub_init_id)
    
    if 'negeri_entiti' in df_right.columns:
        df_right['negeri_entiti'] = df_right['negeri_entiti'].astype(str)
        for data in df_right['negeri_entiti'].unique():
            if data.upper() not in state:
                MapTableState.objects.get_or_create(state=data.upper(), 
                                                    std_state=data.upper(), 
                                                    sub_init_id=sub_init_id)
    
    if 'status_bumiputera_entiti' in df_right.columns:
        df_right['status_bumiputera_entiti'] = df_right['status_bumiputera_entiti'].astype(str)
        for data in df_right['status_bumiputera_entiti'].unique():
            if data.upper() not in bumi_status:
                MapTableBumiStatus.objects.get_or_create(bumi_status=data.upper(), 
                                                         std_bumi_status=data.upper(), 
                                                         sub_init_id=sub_init_id)
    
    if 'sektor_entiti' in df_right.columns:
        df_right['sektor_entiti'] = df_right['sektor_entiti'].astype(str)
        for data in df_right['sektor_entiti'].unique():
            if data.upper() not in biz_sector:
                MapTableBizSector.objects.get_or_create(biz_sector=data.upper(), 
                                                        std_biz_sector=data.upper(), 
                                                        sub_init_id=sub_init_id)
    
    if 'jenis_entiti' in df_right.columns:
        df_right['jenis_entiti'] = df_right['jenis_entiti'].astype(str)
        for data in df_right['jenis_entiti'].unique():
            if data.upper() not in biz_ownership_type:
                MapTableBizOwnershipType.objects.get_or_create(biz_ownership_type=data.upper(), 
                                                               std_biz_ownership_type=data.upper(), 
                                                               sub_init_id=sub_init_id)
    df = df.replace('nan', '')
    df_wrong = df_wrong.replace('nan', '')
    df_right = df_right.replace('nan', '')
    
    return dict_gender, dict_race, dict_state, df, df_wrong, df_right


def move_file_temp(template_name, file_name, sub_init_id):
    """_summary_
    Move /temp to /transition

    Args:
        template_name (_type_): _description_
        file_name (_type_): _description_
        sub_init_id (_type_): _description_
    """
    nextmonth = dt.date.today() + relativedelta.relativedelta(months=1)
    this_day = datetime.today()
    rp_date = str(this_day.year) + '-' + str(this_day.month) + '-' + '16'
    next_rp_date = str(nextmonth.year) + '-' + str(nextmonth.month) + '-' + '16'
    rp_date = datetime.strptime(rp_date, '%Y-%m-%d')
    next_rp_date = datetime.strptime(next_rp_date, '%Y-%m-%d')
    
    if this_day <= rp_date:
        today = rp_date.date() - timedelta(days=1)
    else:
        today = next_rp_date.date() - timedelta(days=1)
        
    try:
        os.makedirs("{}/transition/{}/{}".format(mainPath, template_name, today))
    except Exception as e:
        logging.exception("Exception occured")
        
    now = datetime.now()
    src = "{}/temp/{}.csv".format(mainPath, file_name)
    dest = "{}/transition/{}/{}/{}_{}.csv".format(mainPath, template_name, today, sub_init_id, file_name)
    new_dest = "{}/transition/{}/{}/{}_{}_{}.csv".format(mainPath, template_name, 
                                                         today, sub_init_id, file_name, 
                                                         now.strftime("%d%m%Y_%H%M%S"))
    copy_file_to_target(src, dest, new_dest)
    
    
def move_download_rejected(file_name):
    """_summary_
    Move /temp_download_rejected to /home_download_rejected to render page

    Args:
        file_name (_type_): _description_

    Returns:
        _type_: _description_
    """
    now = datetime.now()
    src = "{}/temp_download_rejected/rejected_rows_{}.csv".format(mainPath, file_name)
    dest = "{}/home_download_rejected/rejected_rows_{}.csv".format(mainPath, file_name)
    new_dest = "{}/home_download_rejected/rejected_rows_{}_{}.csv".format(mainPath, file_name, 
                                                                          now.strftime("%d%m%Y_%H%M%S"))
    rejected_filename = move_file_to_target_rejected(src, dest, new_dest)
    return rejected_filename
    
    
def move_download_accepted(file_name):
    """_summary_
    Move /temp_download_accepted to /home_download_accepted to render page

    Args:
        file_name (_type_): _description_

    Returns:
        _type_: _description_
    """
    now = datetime.now()
    src = "{}/temp_download_accepted/accepted_rows_{}.csv".format(mainPath, file_name)
    dest = "{}/home_download_accepted/accepted_rows_{}.csv".format(mainPath, file_name)
    new_dest = "{}/home_download_accepted/accepted_rows_{}_{}.csv".format(mainPath, file_name, 
                                                                          now.strftime("%d%m%Y_%H%M%S"))
    accepted_filename = move_file_to_target_accepted(src, dest, new_dest)
    return accepted_filename
    

def retrieve_submission_data(information, accepted_filename, rejected_filename):
    """_summary_
    Send submission information to the server

    Args:
        information (_type_): _description_
        accepted_filename (_type_): _description_
        rejected_filename (_type_): _description_
    """
    log = UserLogSubmission(full_name=information['user_full_name'], agency_name=(information['agency_name']),
                          initiative=information['initiative'], filename=information['file_name'], 
                          submission_time=information['date'], sub_init_id=information['sub_init_id'], 
                          accepted_filename=accepted_filename, rejected_filename=rejected_filename, 
                          accepted_rows=information['count_row_accepted'], rejected_rows=information['count_row_rejected'],
                          sum_accepted=information['sumaccepted'], sum_rejected=information['sumrejected'])
    log.save()
    
    
def dq_lvl_one(request):
    """_summary_
    Convert cleaned Dataframe to csv file. Also, move the file from /temp to /transition.
    After converting, move the file to /transition and also email the specific user about the submission

    Args:
        request (_type_): _description_

    Returns:
        _type_: _description_
    """
    if request.method == 'POST':
        file_name = request.POST['file_name']
        initiative = request.POST['initiative']
        sumaccepted = request.POST['sumaccepted']
        sumrejected = request.POST['sumrejected']
        count_row_accepted = request.POST['count_row_accepted']
        count_row_rejected = request.POST['count_row_rejected']
        result_msg = request.POST['result_msg']
        msg = request.POST['msg']
        row_rejected = request.POST['row_rejected']
        row_accepted = request.POST['row_accepted']
        template_name = request.POST['template_name']
        file_ext = request.POST['file_ext']
        sub_init_id = request.POST['sub_init_id']
        
        df_clean_right = usersData(file_name, sub_init_id)[5]
        df_clean_right['submission_time'] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        df_clean_right.to_csv("{}/temp/{}.csv".format(mainPath, file_name), mode='w+', index=False)
        
        init_id_data = MasterlistBudget(subinit_name=initiative, sub_init_id=sub_init_id).pk
        ia_user_list = list(UserProfile.objects.filter(
            init_id_assc=init_id_data, is_lead=False, 
            user__groups__name='ia_group'
            ).exclude(user__email='').values_list('user__email', flat=True))
        ia_lead_list = list(UserProfile.objects.filter(
            init_id_assc=init_id_data, is_lead=True, 
            user__groups__name='ia_group'
            ).exclude(user__email='').values_list('user__email', flat=True))
        pid__user_list = list(UserProfile.objects.filter(
            init_id_assc=init_id_data, is_lead=False, 
            user__groups__name='pid_group'
            ).exclude(user__email='').values_list('user__email', flat=True))
        matching_user = ia_lead_list + ia_user_list + pid__user_list
        
        now = datetime.now()
        information = {'email_IA':matching_user, 'file_name':file_name, 'initiative':initiative, 
                       'sumaccepted':sumaccepted, 'sumrejected':sumrejected, 
                       'count_row_accepted':count_row_accepted, 'count_row_rejected':count_row_rejected, 
                       'date':dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                       'user_full_name':request.user.get_full_name(), 
                       'agency_name':str(request.user.userprofile.agency_name),
                       'result_msg':result_msg, 'msg':msg, 'row_rejected':row_rejected, 
                       'row_accepted':row_accepted, "template_name":template_name, 
                       "file_ext":file_ext, "sumaccepted":sumaccepted, "sumrejected":sumrejected, 
                       "sub_init_id":sub_init_id, 'username': str(request.user.username),
                        }
        
        accepted_filename = move_download_accepted(file_name)
        rejected_filename = move_download_rejected(file_name)
        retrieve_submission_data(information, accepted_filename, rejected_filename)
        move_file_temp(template_name, file_name, sub_init_id)
        logger.info("Submit file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                    str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Submit file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                                   str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        notify_IAs(information)
        return redirect('/')
    
    
def download_file_accepted(request):
    """_summary_
    Download accepted rows

    Args:
        request (_type_): _description_

    Returns:
        _type_: _description_
    """
    if request.method == "POST" :
        file_name = request.POST['file_name']
        full_file_name = "{}/temp_download_accepted/{}.csv".format(mainPath, file_name)
        file_name_w_ext = "{}.csv".format(file_name)

        fl = open(full_file_name, encoding='UTF-8')
        mime_type, _ = mimetypes.guess_type(full_file_name)
        response = HttpResponse(fl, content_type=mime_type)
        response['Content-Disposition'] = "attachment; filename=%s" % file_name_w_ext
        logger.info("Download accepted file " + file_name_w_ext + " - " + str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Download accepted file " + file_name_w_ext + " - " + 
                                   str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        return response
    
    
def download_file_rejected(request):
    """_summary_
    Download rejected rows

    Args:
        request (_type_): _description_

    Returns:
        _type_: _description_
    """
    if request.method == "POST" :
        file_name = request.POST['file_name']
        full_file_name = "{}/temp_download_rejected/{}.csv".format(mainPath, file_name)
        file_name_w_ext = "{}.csv".format(file_name)

        fl = open(full_file_name, encoding='UTF-8')
        mime_type, _ = mimetypes.guess_type(full_file_name)
        response = HttpResponse(fl, content_type=mime_type)
        response['Content-Disposition'] = "attachment; filename=%s" % file_name_w_ext
        logger.info("Download rejected file " + file_name + " - " + str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Download rejected file " + file_name_w_ext + " - " + 
                                   str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        return response


def countX(lst, x):
    """_summary_
    Count empty columns/rows in a list 

    Args:
        lst (_type_): _description_
        x (_type_): _description_

    Returns:
        _type_: _description_
    """
    count = 0
    for ele in lst:
        if (ele == x):
            count = count + 1
    return count


def create_merged_cell_lookup(sheet):
    """_summary_
    The key-value pairs (dict) of merged cell and top value

    Args:
        sheet (_type_): _description_

    Returns:
        dict: _description_
    """
    merged_lookup = {}
    for cell_group in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        if min_col == max_col:
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            merged_lookup[str(cell_group)] = top_left_cell_value
    return merged_lookup


def check_ext(file_ext, file_name, sheet_name, init_id, suffix):
    """_summary_
    Check extension of the file. Different file ext has different process

    Args:
        file_ext (_type_): _description_
        file_name (_type_): _description_
        sheet_name (_type_): _description_
        init_id (_type_): _description_
        suffix (_type_): _description_

    Returns:
        _type_: _description_
    """
    if file_ext in ('xlsx', 'xls'):
        wb = load_workbook("{}/temp/{}.xlsx".format(mainPath, file_name))
        if sheet_name in wb.sheetnames:
            workingFile = convert_to_csv(file_name, file_ext, sheet_name, init_id, suffix)
        else:
            workingFile = "INVALID_FILE"
    elif file_ext in ('csv', 'txt'):
        workingFile = convert_to_csv(file_name, file_ext, sheet_name, init_id, suffix)
    lowercase_header(workingFile)
    return workingFile
    
    
def rem_missing_rows(request, workingFile):
    """_summary_
    Remove missing rows

    Args:
        request (_type_): _description_
        workingFile (_type_): _description_

    Returns:
        _type_: _description_
    """
    if request.method == 'POST':
        df = pd.read_csv(workingFile)
        df = df.dropna(axis=0, how='all')
        x = True    
        dfcolumn = df.columns
        checkHeader = dfcolumn.str.contains('Unnamed', na=False)
        countBlank = countX(checkHeader.tolist(), x)
        content = {'countBlank': countBlank, 'dfcolumn': dfcolumn}
        return content
    
    
def status_one(request, file_name, initiative, template_name, file_ext, sub_init_id):
    """_summary_
    Return status one (Less columns than file template)

    Args:
        request (_type_): _description_
        file_name (_type_): _description_
        initiative (_type_): _description_
        template_name (_type_): _description_
        file_ext (_type_): _description_
        sub_init_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    remainingHeaders = list(set(templateColumn(initiative)[1]).difference(procSpecificCol(file_name, sub_init_id)['list_col_name']))
    messages.error(request, "Fail yang dimuat naik tidak mengandungi kolum yang berikut. Sila semak dan muat naik semula fail anda")
    for header in range (len(remainingHeaders)):
        messages.error(request, remainingHeaders[header]) 
    context = {
            "dataRemaining":remainingHeaders,
            "file_name":file_name,
            "countTemplate":templateColumn(initiative)[0],
            "countUser":procSpecificCol(file_name, sub_init_id)['countCol'],
            "initiative":initiative,
            "template_name":template_name,
            "file_ext":file_ext,
            "sub_init_id":sub_init_id,
            }
    return render(request, 'upload/check-column.html', context)


def status_two(request, file_name, initiative, template_name, file_ext, sub_init_id):
    """_summary_
    Return status two (More column than file template)

    Args:
        request (_type_): _description_
        file_name (_type_): _description_
        initiative (_type_): _description_
        template_name (_type_): _description_
        file_ext (_type_): _description_
        sub_init_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    remainingHeaders = list(set(procSpecificCol(file_name, sub_init_id)['list_col_name']).difference(templateColumn(initiative)[1]))
    remainingHeaders_2 = list(set(templateColumn(initiative)[1]).difference(procSpecificCol(file_name, sub_init_id)['list_col_name']))
    messages.error(request, "Fail yang dimuat naik mengandungi lebihan kolum berikut. Ia mengandungi " + 
                   str(len(remainingHeaders)) + " kolum yang tidak dikenal pasti")
    for header in range (len(remainingHeaders)):
        messages.error(request, remainingHeaders[header]) 
    context = {
        "dataUser":procSpecificCol(file_name, sub_init_id)['list_col_name'], 
        "dataTemplate":templateColumn(initiative)[1], 
        "dataRemaining":remainingHeaders, 
        "dataRemaining2":remainingHeaders_2, 
        "file_name":file_name,
        "countTemplate":templateColumn(initiative)[0],
        "countUser":procSpecificCol(file_name, sub_init_id)['countCol'],
        "initiative":initiative,
        "template_name":template_name,
        "file_ext":file_ext,
        "sub_init_id":sub_init_id,
        }
    return render(request, 'upload/check-column.html', context)


def status_three(request, file_name, initiative, template_name, file_ext, sub_init_id, typeField):
    """_summary_
    Return status three for ALL data templates except financing & summary (a&b)

    Args:
        request (_type_): _description_ 
        file_name (_type_): _description_
        initiative (_type_): _description_
        template_name (_type_): _description_
        file_ext (_type_): _description_
        sub_init_id (_type_): _description_

    Returns:
        _type_: _description_
    """

    if ((icValidation(procSpecificCol(file_name, sub_init_id)['icField'], file_name, typeField)[0] == "Success" or 
         brnValidation(procSpecificCol(file_name, sub_init_id)['brnField'])[0] == "Success") or 
        (dateValidation(procSpecificCol(file_name, sub_init_id)['dateField'])[0] == "Success")):
        result_msg = "Success"
        context = {
            "file_name":file_name,
            "column_name":procSpecificCol(file_name, sub_init_id)['list_col_name'],
            "dataIC":procSpecificCol(file_name, sub_init_id)['icField'], 
            "dataUser":procSpecificCol(file_name, sub_init_id)['dataset'], 
            "resultwrong":procSpecificCol(file_name, sub_init_id)['resultwrong_show'],
            "resultright":procSpecificCol(file_name, sub_init_id)['resultright_show'],
            "sumaccepted":procSpecificCol(file_name, sub_init_id)['total_accepted_disb'],
            "sumrejected":procSpecificCol(file_name, sub_init_id)['total_rejected_disb'],
            "row_rejected":procSpecificCol(file_name, sub_init_id)['resultwrong_filename'],
            "row_accepted":procSpecificCol(file_name, sub_init_id)['resultright_filename'],
            "sum_bene_accepted":procSpecificCol(file_name, sub_init_id)['count_right'],
            "sum_bene_rejected":procSpecificCol(file_name, sub_init_id)['count_wrong'],
            "count_row_rejected":procSpecificCol(file_name, sub_init_id)['count_row_rejected'],
            "count_row_accepted":procSpecificCol(file_name, sub_init_id)['count_row_accepted'],
            "initiative":initiative,
            "result_msg":result_msg,
            "template_name":template_name,
            "file_ext":file_ext,
            "sub_init_id":sub_init_id,
        }
        logger.info("Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                    str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Successfully upload file " + file_name + ".csv - Initiative: " + 
                                   str(initiative) + " - " + str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        return render(request, 'upload/check-data.html', context)
    else:
        result_msg = "Failed"
        msg = "Ralat telah dikenal pasti pada kolum tarikh/no_id_ind/no_brn_entiti"
        context = {
            "file_name":file_name,
            "column_name":procSpecificCol(file_name, sub_init_id)['list_col_name'],
            "dataIC":procSpecificCol(file_name, sub_init_id)['icField'], 
            "dataUser":procSpecificCol(file_name, sub_init_id)['dataset'], 
            "resultwrong":procSpecificCol(file_name, sub_init_id)['resultwrong_show'],
            "resultright":procSpecificCol(file_name, sub_init_id)['resultright_show'],
            "sumaccepted":procSpecificCol(file_name, sub_init_id)['total_accepted_disb'],
            "sumrejected":procSpecificCol(file_name, sub_init_id)['total_rejected_disb'],
            "row_rejected":procSpecificCol(file_name, sub_init_id)['resultwrong_filename'],
            "row_accepted":procSpecificCol(file_name, sub_init_id)['resultright_filename'],
            "sum_bene_accepted":procSpecificCol(file_name, sub_init_id)['count_right'],
            "sum_bene_rejected":procSpecificCol(file_name, sub_init_id)['count_wrong'],
            "count_row_rejected":procSpecificCol(file_name, sub_init_id)['count_row_rejected'],
            "count_row_accepted":procSpecificCol(file_name, sub_init_id)['count_row_accepted'],
            "initiative":initiative,
            "result_msg":result_msg,
            "msg":msg,
            "template_name":template_name,
            "file_ext":file_ext,
            "sub_init_id":sub_init_id,
        }
        logger.info("Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                    str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Successfully upload file " + file_name + ".csv - Initiative: " + 
                                   str(initiative) + " - " + str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        return render(request, 'upload/check-data.html', context)
    
    
def status_three_no_ic(request, file_name, initiative, template_name, file_ext, sub_init_id):
    """_summary_
    Return status three for financing data template (NO IC)

    Args:
        request (_type_): _description_
        file_name (_type_): _description_
        initiative (_type_): _description_
        template_name (_type_): _description_
        file_ext (_type_): _description_
        sub_init_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    if ((brnValidation(procSpecificCol(file_name, sub_init_id)['brnField'])[0] == "Success") or 
        (dateValidation(procSpecificCol(file_name, sub_init_id)['dateField'])[0] == "Success")):
        result_msg = "Success"
        context = {
            "file_name":file_name,
            "column_name":procSpecificCol(file_name, sub_init_id)['list_col_name'],
            "dataIC":procSpecificCol(file_name, sub_init_id)['brnField'], 
            "dataUser":procSpecificCol(file_name, sub_init_id)['dataset'], 
            "resultwrong":procSpecificCol(file_name, sub_init_id)['resultwrong_show'],
            "resultright":procSpecificCol(file_name, sub_init_id)['resultright_show'],
            "sumaccepted":procSpecificCol(file_name, sub_init_id)['total_accepted_disb'],
            "sumrejected":procSpecificCol(file_name, sub_init_id)['total_rejected_disb'],
            "row_rejected":procSpecificCol(file_name, sub_init_id)['resultwrong_filename'],
            "row_accepted":procSpecificCol(file_name, sub_init_id)['resultright_filename'],
            "sum_bene_accepted":procSpecificCol(file_name, sub_init_id)['count_right'],
            "sum_bene_rejected":procSpecificCol(file_name, sub_init_id)['count_wrong'],
            "count_row_rejected":procSpecificCol(file_name, sub_init_id)['count_row_rejected'],
            "count_row_accepted":procSpecificCol(file_name, sub_init_id)['count_row_accepted'],
            "initiative":initiative,
            "result_msg":result_msg,
            "template_name":template_name,
            "file_ext":file_ext,
            "sub_init_id":sub_init_id,
        }
        logger.info("Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                    str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                                   str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        return render(request, 'upload/check-data.html', context)
    
    else:
        result_msg = "Failed"
        msg = "Ralat telah dikenal pasti pada kolum tarikh/no_id_ind/no_brn_entiti"
        context = {
            "file_name":file_name,
            "column_name":procSpecificCol(file_name, sub_init_id)['list_col_name'],
            "dataIC":procSpecificCol(file_name, sub_init_id)['brnField'], 
            "dataUser":procSpecificCol(file_name, sub_init_id)['dataset'], 
            "resultwrong":procSpecificCol(file_name, sub_init_id)['resultwrong_show'],
            "resultright":procSpecificCol(file_name, sub_init_id)['resultright_show'],
            "sumaccepted":procSpecificCol(file_name, sub_init_id)['total_accepted_disb'],
            "sumrejected":procSpecificCol(file_name, sub_init_id)['total_rejected_disb'],
            "row_rejected":procSpecificCol(file_name, sub_init_id)['resultwrong_filename'],
            "row_accepted":procSpecificCol(file_name, sub_init_id)['resultright_filename'],
            "sum_bene_accepted":procSpecificCol(file_name, sub_init_id)['count_right'],
            "sum_bene_rejected":procSpecificCol(file_name, sub_init_id)['count_wrong'],
            "count_row_rejected":procSpecificCol(file_name, sub_init_id)['count_row_rejected'],
            "count_row_accepted":procSpecificCol(file_name, sub_init_id)['count_row_accepted'],
            "initiative":initiative,
            "result_msg":result_msg,
            "msg":msg,
            "template_name":template_name,
            "file_ext":file_ext,
            "sub_init_id":sub_init_id,
        }
        logger.info("Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                    str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                                   str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        return render(request, 'upload/check-data.html', context)
    
    
def status_three_summary(request, file_name, initiative, template_name, file_ext, sub_init_id):
    """_summary_
    Return status three for summary data template (NO IC & BRN)

    Args:
        request (_type_): _description_
        file_name (_type_): _description_
        initiative (_type_): _description_
        template_name (_type_): _description_
        file_ext (_type_): _description_
        sub_init_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    if (dateValidation(procSpecificCol(file_name, sub_init_id)['dateField'])[0] == "Success"):
        result_msg = "Success"
        context = {
            "file_name":file_name,
            "column_name":procSpecificCol(file_name, sub_init_id)['list_col_name'],
            "dataIC":procSpecificCol(file_name, sub_init_id)['brnField'], 
            "dataUser":procSpecificCol(file_name, sub_init_id)['dataset'], 
            "resultwrong":procSpecificCol(file_name, sub_init_id)['resultwrong_show'],
            "resultright":procSpecificCol(file_name, sub_init_id)['resultright_show'],
            "sumaccepted":procSpecificCol(file_name, sub_init_id)['total_accepted_disb'],
            "sumrejected":procSpecificCol(file_name, sub_init_id)['total_rejected_disb'],
            "row_rejected":procSpecificCol(file_name, sub_init_id)['resultwrong_filename'],
            "row_accepted":procSpecificCol(file_name, sub_init_id)['resultright_filename'],
            "sum_bene_accepted":procSpecificCol(file_name, sub_init_id)['count_right'],
            "sum_bene_rejected":procSpecificCol(file_name, sub_init_id)['count_wrong'],
            "count_row_rejected":procSpecificCol(file_name, sub_init_id)['count_row_rejected'],
            "count_row_accepted":procSpecificCol(file_name, sub_init_id)['count_row_accepted'],
            "initiative":initiative,
            "result_msg":result_msg,
            "template_name":template_name,
            "file_ext":file_ext,
            "sub_init_id":sub_init_id,
        }
        logger.info("Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                    str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                                   str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        return render(request, 'upload/check-data.html', context)
    
    else:
        result_msg = "Failed"
        msg = "Ralat telah dikenal pasti pada kolum tarikh"
        context = {
            "file_name":file_name,
            "column_name":procSpecificCol(file_name, sub_init_id)['list_col_name'],
            "dataIC":procSpecificCol(file_name, sub_init_id)['brnField'], 
            "dataUser":procSpecificCol(file_name, sub_init_id)['dataset'], 
            "resultwrong":procSpecificCol(file_name, sub_init_id)['resultwrong_show'],
            "resultright":procSpecificCol(file_name, sub_init_id)['resultright_show'],
            "sumaccepted":procSpecificCol(file_name, sub_init_id)['total_accepted_disb'],
            "sumrejected":procSpecificCol(file_name, sub_init_id)['total_rejected_disb'],
            "row_rejected":procSpecificCol(file_name, sub_init_id)['resultwrong_filename'],
            "row_accepted":procSpecificCol(file_name, sub_init_id)['resultright_filename'],
            "sum_bene_accepted":procSpecificCol(file_name, sub_init_id)['count_right'],
            "sum_bene_rejected":procSpecificCol(file_name, sub_init_id)['count_wrong'],
            "count_row_rejected":procSpecificCol(file_name, sub_init_id)['count_row_rejected'],
            "count_row_accepted":procSpecificCol(file_name, sub_init_id)['count_row_accepted'],
            "initiative":initiative,
            "result_msg":result_msg,
            "msg":msg,
            "template_name":template_name,
            "file_ext":file_ext,
            "sub_init_id":sub_init_id,
        }
        logger.info("Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                    str(request.user.username))
        try:
            packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', 
                                   "Successfully upload file " + file_name + ".csv - Initiative: " + str(initiative) + " - " + 
                                   str(request.user.username))]
            zabbix_sender.send(packet)
        except:
            pass
        return render(request, 'upload/check-data.html', context)


def dataValidation_submission(request, initiative, file_name, template_name, file_ext, sub_init_id):
    """_summary_
    Data validation after column checking (Required to go through column checking page first
    and then proceed to data validation page)

    Args:
        request (_type_): _description_
        initiative (_type_): _description_
        file_name (_type_): _description_
        template_name (_type_): _description_
        file_ext (_type_): _description_
        sub_init_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    dataframe = pd.read_csv("{}/temp/{}.csv".format(mainPath, file_name))
    
    if check_mand_col(dataframe, initiative) == 'Passed':
        pass
    
    elif check_mand_col(dataframe, initiative) == 'Failed Financing':
        messages.error(request, "Sila sediakan nama yang betul untuk kolum tarikh dan no_brn_entiti")
        return redirect('/upload')
    
    elif check_mand_col(dataframe, initiative) == 'Failed Detail':
        messages.error(request, "Sila sediakan nama yang betul untuk kolum tarikh, no_id_ind dan no_brn_entiti")
        return redirect('/upload')
    
    else:
        messages.error(request, "Sila sediakan nama yang betul untuk kolum tarikh")
        return redirect('/upload')
                    
    remainingHeaders = list(set(procSpecificCol(file_name, sub_init_id)['list_col_name']).difference(templateColumn(initiative)[1]))
    correctionData = pd.read_csv(r"{}/temp/{}.csv".format(mainPath, file_name))
    correctionData = correctionData.drop(columns = remainingHeaders)
    correctionData.to_csv("{}/temp/{}.csv".format(mainPath, file_name), sep=',', encoding='utf-8', index = False)
    
    col_size = correctionData.shape[1]
    if col_size <  templateColumn(initiative)[0]:
        messages.error(request, "Jumlah bilangan kolum anda kurang daripada jumlah bilangan kolum piawaian yang ditetapkan untuk inisiatif " + initiative)
        return redirect('/upload')
    else:
        if 'no_brn_entiti' in dataframe.columns:
            if 'no_id_ind' in dataframe.columns:
                typeField = dataframe.filter(["jenis_id_ind"])
                return status_three(request, file_name, initiative, template_name, file_ext, sub_init_id, typeField)
            else:
                return status_three_no_ic(request, file_name, initiative, template_name, file_ext, sub_init_id)
        else:
            return status_three_summary(request, file_name, initiative, template_name, file_ext, sub_init_id)


@method_decorator(decorators_all, name='dispatch')  
class FileUploadView(View):
    """_summary_
    Class for file upload (Go through all processes until the end)

    Args:
        View (_type_): _description_

    Returns:
        _type_: _description_
    """    
    form_class = FileForm
    template = 'upload/upload.html'

    def get(self, request):
        form = self.form_class(user=request.user)
        return render(request, self.template, {'form':form})

    def post(self, request):
        form = self.form_class(request.user, request.POST, request.FILES)
        file_name_w_ext = str(request.FILES['Fail'])
        file_name = file_name_w_ext.rsplit(".", 1)[0]
        
        if form.is_valid():
            if file_name_w_ext != '':
                file_ext = file_name_w_ext.split('.')[1]
                
                if (file_ext not in ['xlsx', 'xls', 'csv', 'txt']):
                    messages.error(request, "Format fail tidak diterima")
                    return render(request, 'upload/upload.html', {'form':form})
                else:
                    form.save()
                    initiative = form.cleaned_data['Inisiatif']
                    sheet_name = form.cleaned_data['Helaian']
                    
                    template_name = FileTemplate.objects.get(masterlistbudget__subinit_name=initiative)
                    init_assc = list(UserProfile.objects.filter(
                        user__username=self.request.user
                        ).values_list('init_id_assc', flat=True))

                    sub_init_id = MasterlistBudget.objects.filter(
                        id__in=init_assc, subinit_name=initiative
                        ).values_list('sub_init_id', flat=True).first()

                    initID = MasterlistBudget.objects.filter(
                        id__in=init_assc, subinit_name=initiative
                        ).values_list('init_id', flat=True).first()
                        
                    file_name = get_valid_filename(file_name)
                    logger.info("Trying to upload file " + file_name_w_ext + " - Initiative: " + str(initiative) + " - " + str(request.user.username))
                    try:
                        packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', "Trying to upload file " + file_name_w_ext + " - Initiative: " + str(initiative) + " - " + str(request.user.username))]
                        zabbix_sender.send(packet)
                    except:
                        pass
                    
                    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
                    
                    if (file_ext in ['xlsx', 'xls']):
                        read_file = pd.read_excel("{}/temp/{}.{}".format(mainPath, file_name, file_ext), sheet_name)
                        if read_file.empty:
                            messages.error(request, "Tiada data dalam nama helaian " + sheet_name)
                            return redirect('/upload')
                        else:
                            pass
                    
                    try:
                        os.makedirs("{}/ori/{}".format(mainPath, initID))
                    except Exception as e:
                        logging.exception("Exception occured")
                        
                    workingFile = check_ext(file_ext, file_name, sheet_name, initID, suffix)
                    if workingFile == "INVALID_FILE":
                        messages.error(request, "Sila betulkan nama helaian anda")
                        return redirect('/upload')
                    content = rem_missing_rows(request, workingFile)
                    
                    workingFile = check_additional_col(workingFile, initiative)
            
                    if content['countBlank'] > len(content['dfcolumn']) / 2:
                        del_row_col(file_name)
                        dataframe = pd.read_csv("{}/temp/{}.csv".format(mainPath, file_name))
                        
                        if check_mand_col(dataframe, initiative) == 'Passed':
                            pass
                        
                        elif check_mand_col(dataframe, initiative) == 'Failed Financing':
                            messages.error(request, "Sila sediakan nama yang betul untuk kolum tarikh dan no_brn_entiti")
                            return redirect('/upload')
                        
                        elif check_mand_col(dataframe, initiative) == 'Failed Detail':
                            messages.error(request, "Sila sediakan nama yang betul untuk kolum tarikh, no_id_ind dan no_brn_entiti")
                            return redirect('/upload')
                        
                        else:
                            messages.error(request, "Sila sediakan nama yang betul untuk kolum tarikh")
                            return redirect('/upload')
                        
                        if check_init_id(sub_init_id, workingFile) == True:
                            messages.error(request, "Sila betulkan dan isi indeks data anda pada kolum indeks_data")
                            return redirect('/upload')
                        else:
                            pass

                        if checkColumn(workingFile, initiative)[0] == '1':
                            return status_one(request, file_name, initiative, template_name, file_ext, sub_init_id)
                        
                        elif checkColumn(workingFile, initiative)[0] == '2':
                            return status_two(request, file_name, initiative, template_name, file_ext, sub_init_id) 
                        
                        elif checkColumn(workingFile, initiative)[0] == '3':
                            if 'no_brn_entiti' in dataframe.columns:
                                typeField = dataframe.filter(["jenis_id_ind"])
                                if 'no_id_ind' in dataframe.columns:
                                    dataframe = pd.read_csv("{}/temp/{}.csv".format(mainPath, file_name))
                                    return status_three(request, file_name, initiative, template_name, file_ext, sub_init_id, typeField)
                                else:
                                    return status_three_no_ic(request, file_name, initiative, template_name, file_ext, sub_init_id)
                            else:
                                return status_three_summary(request, file_name, initiative, template_name, file_ext, sub_init_id)                            
                    else:
                        dataframe = pd.read_csv("{}/temp/{}.csv".format(mainPath, file_name))
                        if check_mand_col(dataframe, initiative) == 'Passed':
                            pass
                        
                        elif check_mand_col(dataframe, initiative) == 'Failed Financing':
                            messages.error(request, "Sila sediakan nama yang betul untuk kolum tarikh dan no_brn_entiti")
                            return redirect('/upload')
                        
                        elif check_mand_col(dataframe, initiative) == 'Failed Detail':
                            messages.error(request, "Sila sediakan nama yang betul untuk kolum tarikh, no_id_ind dan no_brn_entiti")
                            return redirect('/upload')
                        
                        else:
                            messages.error(request, "Sila sediakan nama yang betul untuk kolum tarikh")
                            return redirect('/upload')
                        
                        if check_init_id(sub_init_id, workingFile) == True:
                            messages.error(request, "Sila betulkan dan isi indeks data anda pada kolum indeks_data")
                            return redirect('/upload')
                        else:
                            pass
                        if checkColumn(workingFile, initiative)[0] == '1':
                            print("case1")
                            return status_one(request, file_name, initiative, template_name, file_ext, sub_init_id)
                        
                        elif checkColumn(workingFile, initiative)[0] == '2':
                            return status_two(request, file_name, initiative, template_name, file_ext, sub_init_id)
                        
                        elif checkColumn(workingFile, initiative)[0] == '3': 
                            if 'no_brn_entiti' in dataframe.columns:
                                typeField = dataframe.filter(["jenis_id_ind"])
                                if 'no_id_ind' in dataframe.columns:
                                    dataframe = pd.read_csv("{}/temp/{}.csv".format(mainPath, file_name))
                                    return status_three(request, file_name, initiative, template_name, file_ext, sub_init_id, typeField)
                                else:
                                    return status_three_no_ic(request, file_name, initiative, template_name, file_ext, sub_init_id)
                            else:
                                return status_three_summary(request, file_name, initiative, template_name, file_ext, sub_init_id)
        else:
            form = self.form_class(user=request.user)
            messages.error(request, "Format fail tidak diterima")
            logger.info("Failed upload file " + file_name_w_ext + " - Initiative: " + str(initiative) + " - " + str(request.user.username))
            try:
                packet = [ZabbixMetric('CentOS PODS Dev Master', 'django.log.user_activity', "Failed upload file " + file_name_w_ext + " - Initiative: " + str(initiative) + " - " + str(request.user.username))]
                zabbix_sender.send(packet)
            except:
                pass
            return render(request, self.template, {'form':form})